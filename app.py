# -*- coding: utf-8 -*-
"""
Created on Thu Aug 24 20:49:44 2023

@author: azzed
"""

## input this directly into the console: pip install xlrd


import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
import base64
import xlrd
from io import BytesIO

# text font and colour
font = """
<style>
.stApp {{
  font-family: 'Helvetica';
 # color: white !important;
}}
</style>
"""
st.markdown(font, unsafe_allow_html=True)


# background of web app:
with open("boats_background.jpg", "rb") as image_file:
    base64_image = base64.b64encode(image_file.read()).decode()

# Create a background style using the base64 string
background_style = f"""
<style>
.stApp {{
  background-image: url("data:image/jpg;base64,{base64_image}");
  background-size: cover;
  background-repeat: no-repeat;
  background-attachment: fixed;
}}
</style>
"""

# Add the background style to the Streamlit app
st.markdown(background_style, unsafe_allow_html=True)


# Streamlit app title
st.title("North Sea Activity Plotter")

# Verify the current working directory
# print(os.getcwd())


def find_adjacent_blocks(df, block):
    # Locate the block in the grid
    loc = df[df == block].stack().index.tolist()
    if not loc:
        return None

    row, col = loc[0]

    # Extract a 5x5 grid around the block using slicing (and handle edge cases)
    sub_grid = df.iloc[max(row-2, 0):min(row+3, df.shape[0]),
                       max(col-2, 0):min(col+3, df.shape[1])]

    return sub_grid

def main():
    # Read Excel file
    df = pd.read_excel('north_sea_blocks.xlsx', header=None, engine='openpyxl')
    #print(f"df after reading north sea blocks has {df.shape[0]} rows")
    # Convert all zero values to blanks
    df = df.applymap(lambda x: '' if str(x) == '0' or x == 0 or x == 0.0 else x)
    #print(f"df after removing zeros has {df.shape[0]} rows")
    # Input block number
    block_number = st.text_input("Enter main block number (please leave out any letters): ")
    #print(f"block number entered was {block_number}")
    if block_number:
       #block_number = int(block_number)
       result = find_adjacent_blocks(df, block_number)
       #print(f"adjacent blocks file has {result.shape[0]} rows")
       if result is None:
           #print("there was no result")
           st.write(f"Block {block_number} not found in the grid.\nPleave leave out any zeros before any single digit, e.g. 21/5")
       else:
           return result
           #print(f"result is {result}")
    #else:
        #print("block number apparently false")


if __name__ == "__main__":
    adjacent_blocks_df = main()
    #print(f"adjacent blocks df, after processing via main function, has {adjacent_blocks_df.shape[0]} rows")
    if adjacent_blocks_df is not None:
        # Proceed with the code to handle uploaded_file and other logic
        uploaded_file = st.file_uploader("Upload the PETS Application Data, after downloading the excel file from here: https://itportal.beis.gov.uk/eng/fox/beis/PETS_EXTERNAL_PUBLICATION/main", type=['xlsx', 'xls'])

        # Read the entire workbook into memory
        if uploaded_file:
            # Read the Excel file into a pandas ExcelFile object
            excel_file = pd.ExcelFile(uploaded_file)
   
            # Read each sheet into a DataFrame
            application_data_drilling = pd.read_excel(excel_file, 'Drilling', header=1)
            application_data_pipeline = pd.read_excel(excel_file, 'Pipeline', header=1)
            application_data_well_intervention = pd.read_excel(excel_file, 'Well Intervention', header=1)
            application_data_decommissioning = pd.read_excel(excel_file, 'Decommissioning', header=1)
            application_data_standalone = pd.read_excel(excel_file, 'Standalone', header=1)
            


#print(f"application data drilling columns: {application_data_drilling.columns}")


# Function to find matches for the Drilling sheet
def find_matching_rows_drilling(block):
    matching_rows = set() # Using a set to ensure uniqueness
    for index, row in application_data_drilling.iterrows():
        if str(block) in str(row['Quadrant/Block']):
            app_type = row['Application Type'].lower()
            formatted_app_type = None
            
            if 'drilling' in app_type:
                formatted_app_type = 'drilling'
            elif 'consent to locate' in app_type:
                formatted_app_type = 'consent to locate'
            elif 'seismic' in app_type:
                formatted_app_type = 'seismic'
            elif 'sub-bottom' in app_type:
                formatted_app_type = 'sub-bottom profiler'
            elif 'marine' in app_type:
                formatted_app_type = 'marine survey'

            if formatted_app_type:
                formatted_str = f"{row['Operator']} ({row['Field/Prospect']})/{row['MoDU']}/{formatted_app_type}"
                matching_rows.add(formatted_str)
    
    return '\n'.join(matching_rows)

# Function to find matches for the Pipeline sheet
def find_matching_rows_pipeline(block):
    matching_rows = set()
    for index, row in application_data_pipeline.iterrows():
        if str(block) in str(row['Start Quadrant/Block']) or str(block) in str(row['End Quadrant/Block']):
            app_type = row['Application Type'].lower()
            formatted_app_type = None
            
            if 'drilling' in app_type:
                formatted_app_type = 'drilling'
            elif 'consent to locate' in app_type:
                formatted_app_type = 'consent to locate'
            elif 'seismic' in app_type:
                formatted_app_type = 'seismic'
            elif 'sub-bottom' in app_type:
                formatted_app_type = 'sub-bottom profiler'
            elif 'marine' in app_type:
                formatted_app_type = 'marine survey'

            if formatted_app_type:
                formatted_str = f"{row['Operator']}/{row['Main Pipeline Number']}/{formatted_app_type}"
                matching_rows.add(formatted_str)
    
    return '\n'.join(matching_rows)

# Function to find matches for the Well Intervention sheet
def find_matching_rows_well_intervention(block):
    matching_rows = set() # Using a set to ensure uniqueness
    for index, row in application_data_well_intervention.iterrows():
        if str(block) in str(row['Quadrant/Block']):
            app_type = row['Application Type'].lower()
            formatted_app_type = None
            
            if 'drilling' in app_type:
                formatted_app_type = 'drilling'
            elif 'consent to locate' in app_type:
                formatted_app_type = 'consent to locate'
            elif 'seismic' in app_type:
                formatted_app_type = 'seismic'
            elif 'sub-bottom' in app_type:
                formatted_app_type = 'sub-bottom profiler'
            elif 'marine' in app_type:
                formatted_app_type = 'marine survey'

            if formatted_app_type:
                formatted_str = f"{row['Operator']} ({row['Field/Prospect']})/{row['MoDU/Vessel']}/{formatted_app_type}"
                matching_rows.add(formatted_str)
    
    return '\n'.join(matching_rows)

# Function to find matches for the Decommissioning sheet
def find_matching_rows_decom(block):
    matching_rows = set() # Using a set to ensure uniqueness
    for index, row in application_data_decommissioning.iterrows():
        if str(block) in str(row['Quadrant/Block']) and str('Subsea') in str(row['Facility/Installation Type']):
            app_type = row['Application Type'].lower()
            formatted_app_type = None
            
            if 'drilling' in app_type:
                formatted_app_type = 'drilling'
            elif 'consent to locate' in app_type:
                formatted_app_type = 'consent to locate'
            elif 'seismic' in app_type:
                formatted_app_type = 'seismic'
            elif 'sub-bottom' in app_type:
                formatted_app_type = 'sub-bottom profiler'
            elif 'marine' in app_type:
                formatted_app_type = 'marine survey'

            if formatted_app_type:
                formatted_str = f"{row['Operator']} ({row['Name or Identifier of Facility/Installation']})/{formatted_app_type}"
                matching_rows.add(formatted_str)
    
    return '\n'.join(matching_rows)

# Function to find matches for the Standalone sheet
def find_matching_rows_standalone(block):
    matching_rows = set() # Using a set to ensure uniqueness
    for index, row in application_data_standalone.iterrows():
        if str(block) in str(row['Quadrant/Block']):
            app_type = row['Application Type'].lower()
            formatted_app_type = None
            
            if 'drilling' in app_type:
                formatted_app_type = 'drilling'
            elif 'consent to locate' in app_type:
                formatted_app_type = 'consent to locate'
            elif 'seismic' in app_type:
                formatted_app_type = 'seismic'
            elif 'sub-bottom' in app_type:
                formatted_app_type = 'sub-bottom profiler'
            elif 'marine' in app_type:
                formatted_app_type = 'marine survey'

            if formatted_app_type:
                formatted_str = f"{row['Operator']}/{formatted_app_type}"
                matching_rows.add(formatted_str)
    
    return '\n'.join(matching_rows)

# debugging
#print(adjacent_blocks_df)


# Iterate through the adjacent_blocks_df DataFrame and search for matches in all sheets
if adjacent_blocks_df is not None and uploaded_file is not None:
    for row in range(adjacent_blocks_df.shape[0]):
        for col in range(adjacent_blocks_df.shape[1]):
            block_number = adjacent_blocks_df.iloc[row, col]
            # to skip blank rows and avoid matching them:
            if block_number == 0 or block_number == '':
                continue
                
            # Drilling matches
            matching_info_drilling = find_matching_rows_drilling(block_number)
    
            # Pipeline matches
            matching_info_pipeline = find_matching_rows_pipeline(block_number)
            
            # Well intervention matches
            matching_info_well_intervention = find_matching_rows_well_intervention(block_number)
    
            # Well intervention matches
            matching_info_decom = find_matching_rows_decom(block_number)
            
            # Standalone matches
            matching_info_standalone = find_matching_rows_standalone(block_number)
    
            # Combine all results
            combined_value = f"{block_number}"
            if matching_info_drilling:
                combined_value += f"\n{matching_info_drilling}"
            if matching_info_pipeline:
                combined_value += f"\n{matching_info_pipeline}"
            if matching_info_well_intervention:
                combined_value += f"\n{matching_info_well_intervention}"
            if matching_info_decom:
                combined_value += f"\n{matching_info_decom}"
            if matching_info_standalone:
                combined_value += f"\n{matching_info_standalone}"
                
            adjacent_blocks_df.iloc[row, col] = combined_value  # Replace the block number with the combined value

# Save the modified DataFrame to a new Excel file
output_file_name = 'modified_adjacent_blocks.xlsx'
if adjacent_blocks_df is not None and uploaded_file is not None:
    adjacent_blocks_df.to_excel(output_file_name, index=False, header=False, engine='openpyxl')
    
    print("Grid filled out and saved to 'modified_adjacent_blocks.xlsx'.")


    
    # format the output excel sheet
    
    # Load the workbook
    workbook = load_workbook(output_file_name)
    
    # Access the active worksheet
    worksheet = workbook.active
    
    # Create styles for the borders
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    thick_border = Border(left=Side(style='thick'),
                          right=Side(style='thick'),
                          top=Side(style='thick'),
                          bottom=Side(style='thick'))
    
    # Amber fill
    amber_fill = PatternFill(start_color="FFC000",
                             end_color="FFC000", fill_type="solid")
    
    # Light grey fill
    light_grey_fill = PatternFill(start_color="D3D3D3",
                                  end_color="D3D3D3", fill_type="solid")
    
    # Set column widths, row heights, and thin borders for the 5x5 grid
    wrap_alignment = Alignment(wrap_text=True)
    for row in worksheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
        for cell in row:
            worksheet.column_dimensions[cell.column_letter].width = 20
            worksheet.row_dimensions[cell.row].height = 110
            cell.alignment = wrap_alignment
            cell.border = thin_border
            # Apply amber fill if the cell contains the word 'seismic'
            if 'seismic' in str(cell.value).lower():
                cell.fill = amber_fill
            # If the cell is blank, set the value to "out of bounds" and apply light grey fill
            if cell.value is None or str(cell.value).strip() == "" or str(cell.value)=="nan":
                cell.value = "out of bounds"
                cell.fill = light_grey_fill
    
    # Apply thick border to the middle cell
    worksheet['C3'].border = thick_border
    
    # Save the workbook with the formatting
    workbook.save(output_file_name)
    
    
    # Offer download link for the file
    
    def get_table_download_link(file_name):
        with open(file_name, 'rb') as f:
            bytes_data = f.read()
            b64 = base64.b64encode(bytes_data).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}">Click here to download the grid in Excel format</a>'
            return href
    
    # Create the download link
    download_link = get_table_download_link(output_file_name)
    # Write the download link to the Streamlit app
    st.markdown(download_link, unsafe_allow_html=True)
    


